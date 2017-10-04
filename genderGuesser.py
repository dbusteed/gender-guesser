#Davis Busteed -- busteed.davis@gmail.com
#
#September 2017
#
#This is just an adaption of a basic decision tree example
#that i picked up in a youtube video by Siraj Raval
#

#menu
print("\n")
print("-----Gender Guesser-----")
print("\n")
print('Initializing data sets...\n')

#gonna need these
from sklearn import tree
import openpyxl

#prep excel file
wb = openpyxl.load_workbook('genderData.xlsx')
mSheet = wb.get_sheet_by_name('male')
fSheet = wb.get_sheet_by_name('female')

#make lits
values = []
labels = []

#get values/labels for male
for i in range(2, mSheet.max_row + 1):
	h = mSheet.cell(row=i, column=1).value
	w = mSheet.cell(row=i, column=2).value
	values.append([h,w])
	labels.append('male')

#get values/labels for female
for i in range(2, fSheet.max_row + 1):
	h = fSheet.cell(row=i, column=1).value
	w = fSheet.cell(row=i, column=2).value
	values.append([h,w])
	labels.append('female')

#make data tree
clf = tree.DecisionTreeClassifier()

#fit the values to the labels
clf = clf.fit(values, labels)
	

#get inputs from user
h = int(input("What is your height in inches?: "))
w = int(input("What is your weight in pounds?: "))
print("\n")

#create a prediction based of these values (run thru the dec tree
pred = clf.predict([[h, w]])

#print the prediction, and see if it got it right
fb = input("Are you a %s? [y/n] " %tuple(pred))

print("\n")

#got it right?
if fb == 'y':
	print("Awesome!")
else:
	print("Oops!")
	
#ask user if we can add the data to our excel sheet	
print("\n")
choice = input("Do you want to add your data to the decision tree so that future predictions are more accurate?: [y/n] ")

#if es okay, add to the sheet, depending on if the AI was right / their gender
if choice == 'y':
	print("\nThanks, your data is appreciated.")
	
	if fb == 'y':
		if pred == ['male']:
			mSheet.cell(row=(mSheet.max_row + 1), column=1).value = h
			mSheet.cell(row=(mSheet.max_row), column=2).value = w		
		else:
			fSheet.cell(row=(fSheet.max_row + 1), column=1).value = h
			fSheet.cell(row=(fSheet.max_row), column=2).value = w
	else:
		if pred == ['male']:
			fSheet.cell(row=(fSheet.max_row + 1), column=1).value = h
			fSheet.cell(row=(fSheet.max_row), column=2).value = w
		else: 
			mSheet.cell(row=(mSheet.max_row + 1), column=1).value = h
			mSheet.cell(row=(mSheet.max_row), column=2).value = w
	
	#save the data
	wb.save('genderData.xlsx')
	
elif choice == 'n':
	print("\nOkay. Your data won't be used.")
			
		
	
